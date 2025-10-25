import json
import re
from typing import OrderedDict
from dashboard.app import IsDebugEnabled, g_
from flask import make_response, request, jsonify, make_response
from io import BytesIO
from openpyxl import load_workbook
from pymupdf import Document
from pdf2image import convert_from_bytes
from pdf2image.exceptions import PDFInfoNotInstalledError, PDFPageCountError, PDFSyntaxError
from os import getcwd
from tempfile import TemporaryDirectory
from glob import glob
from zipfile import ZipFile, ZIP_DEFLATED
from uuid import uuid4
from os import mkdir, rename
from os.path import basename
from shutil import rmtree
from asyncio import gather
from json import loads
from dotenv import load_dotenv
load_dotenv()
from openai import ChatCompletion, api_key
from openai.error import RateLimitError
from os import getenv
from aiohttp import ClientSession
api_key            = getenv('OPENAI_API_KEY')
OPENROUTER_API_KEY = getenv('OPEN_ROUTER_KEY')
from json import dumps
from asyncio import sleep as async_sleep
from traceback import format_exc
from dashboard.alchemy import AlchemyLogs
from dashboard.app import GetCurrentUserid


@g_.app.route("/dataset-finetunning_read-excel", methods=['POST']) #type: ignore 
def read_excel():
    stream = BytesIO(request.get_data())
    wb     = load_workbook(filename=stream)
    text   = ''
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows():
            for cell in row:
                if cell.value != None:
                    text += str(cell.value) + ' '
            text += '\n'  
        text += '\n'       
    while text.find('\n\n\n') != -1:
        text = text.replace('\n\n\n', '\n\n')    
    return jsonify({'Text': text})

@g_.app.route("/dataset-finetunning_read-pdf", methods=['POST']) #type: ignore 
def read_pdf():    
    stream     = BytesIO(request.get_data())

    doc        = Document(stream=stream)
    text_pages = [page.get_text() + '\n' for page in doc] #type: ignore 
    text       = ''.join(text_pages) 
    has_toc    = False
    has_chaps  = False
    chapters   = []
    chapter    = ''    
    for line in text.splitlines():
        line_low = line.lower()
        if line_low == 'table of contents': has_toc = True
        split = line_low.split(' ')
        if len(split) == 2 and split[1].isalnum() and split[0] == 'chapter': 
            if has_chaps == True:
                chapters.append(chapter)
                chapter = ''
            has_chaps = True
        if has_chaps: chapter += line + '\n'
          
    if has_toc and has_chaps:
        return jsonify({'Text': '', 'Chapters' : chapters})   
    return jsonify({'Text': text, 'Chapters' : []} )  

@g_.app.route("/dataset-finetunning_extract-pdf-create-job", methods=['GET']) #type: ignore   
def pdf_extract_cerate():
    dir = '/tmp/' + str(uuid4())
    mkdir(dir)
    return jsonify({'JobId': dir})

def Get1stPage(stream):
    doc        = Document(stream=stream)
    text_pages = [page.get_text() + '\n' for page in doc] #type: ignore     
    has_toc    = False    
    chapters   = []
    chapter    = ''    
    for index, text in enumerate(text_pages):
        for line in text.splitlines():
            line_low = line.lower()
            if line_low == 'table of contents': has_toc = True
            split = line_low.split(' ')
            if len(split) == 2 and split[1] == '01' and split[0] == 'chapter': 
                return index + 1
    return 0   

def GetChapters(stream)->list:
    doc        = Document(stream=stream)
    text_pages = [page.get_text() + '\n' for page in doc] #type: ignore     
    has_toc    = False    
    chapters   = []
    chapter    = -1    
    index      = 0
    for index, text in enumerate(text_pages):
        for line in text.splitlines():
            line_low = line.lower()
            if line_low == 'table of contents': has_toc = True
            split = line_low.split(' ')
            if len(split) == 2 and split[1].isalnum() and split[0] == 'chapter': 
                if chapter > -1:
                    chapters[-1]['end'] = index
                chapters.append({'start' : index+2, 'end' : 0})
                chapter += 1
    if chapter > -1:  
        chapters[-1]['end'] = index        
    return chapters   

@g_.app.route("/dataset-finetunning_extract-pdf", methods=['POST']) #type: ignore   
def pdf_extract():
    stream       = BytesIO(request.get_data()) 
    article_name = request.args.get('Name').replace('.pdf', '')                                     #type: ignore
    chapters     = GetChapters(stream)
    if len(chapters) == 0:
        convert_from_bytes(stream.read(), fmt='png', output_folder=request.args.get('JobId'))       #type: ignore
    else:
        for chapter in chapters:
            print(chapter)
            convert_from_bytes(BytesIO(request.get_data()).read(), fmt='png', output_folder=request.args.get('JobId'),   #type: ignore
            first_page=chapter['start'], last_page=chapter['end'])                                  #type: ignore   
        
    for file in glob(request.args.get('JobId') + '/*.png'):     #type: ignore                                                        
        f_name = basename(file).split('.png')[0]
        split  = f_name.split('-')
        if split[-1].isalnum():           
            page_number = split[-1]
            new_name = article_name + '_' + page_number + '.png' #type: ignore   
            rename(file, str(request.args.get('JobId')) + '/' + new_name)           
    return jsonify({'Err': 'S_OK'})

@g_.app.route("/dataset-finetunning_extract-pdf-finish-job", methods=['GET']) #type: ignore    
def pdf_extract_images():
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "a", ZIP_DEFLATED, False) as zip_file:
        for file in glob(request.args.get('JobId') + '/*.png'):   #type: ignore                     
            zip_file.write(file, file)                   
    rmtree(request.args.get('JobId'))                             #type: ignore    
    response = make_response(zip_buffer.getvalue())
    response.headers['Content-Type']        = 'application/octet-stream'
    response.headers['Content-Disposition'] = 'inline; filename=Articles'
    return response       

def dump_error(name, e):
    fname = name.split('.')[0] + '.txt'
    with open(fname, 'w') as f:
        f.write(str(e))  
        f.write(format_exc())  

def GetPrompts(data, promtps, format):   
    if data['Type'] == 'text':
        messages = [
            {'role': 'system', 'content': promtps['System'].replace('{FILENAME}', data['Name']).replace('{DATA}', data['Content'])}, 
            {'role' : 'user', 'content': promtps['User'].replace('{FILENAME}', data['Name']).replace('{DATA}', data['Content'])}
        ] 
    elif data['Type'] == 'base64':              
        messages = [
            {'role': 'system', 'content': promtps['System - image'].replace('{FILENAME}', data['Name'])}, 
            {'role' : 'user', 'content': [
                    {
                        'type' : "text",
                        'text' :  promtps['User - image'].replace('{FILENAME}', data['Name']),
                    },
                    {
                        'type'     : "image_url",
                        'image_url': {'url' : data['Content']}
                    }
                ]
            }
        ] 
    else:
        messages = 'Invalid data type'        
    return messages  

async def __GenQandAWithChatGPT(data, promtps, format, retries=0):  
    try:
        messages = GetPrompts(data, promtps, format)

        if IsDebugEnabled():
            return json.loads(f'{{ "{data["Name"]}" : "Placeholder Reply for {data["Name"]}" }}')

        # Prepare your tool
        if format == 'json':
            tools = [{
                "type": "function",
                "function": {
                    "name": "get_questions_and_answers",
                    "description": "enumerate questions and answers",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "QAndA": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "question": {"type": "string"},
                                        "answer":   {"type": "string"},
                                        "context":  {"type": "string"},
                                    }
                                }
                            }
                        }
                    }
                }
            }]
        elif format == 'stf':
            tools = [{
                "type": "function",
                "function": {
                    "name": "get_questions_and_answers",
                    "description": "enumerate questions and answers",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "QAndA": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "text": {"type": "string"}
                                    }
                                }
                            }
                        }
                    }
                }
            }]
        elif format == 'stf-original':
            tools = [{
                "type": "function",
                "function": {
                    "name": "get_questions_and_answers",
                    "description": "enumerate questions and answers",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "QAndA": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "instruction": {"type": "string"},
                                        "input":       {"type": "string"},
                                        "output":      {"type": "string"},
                                    }
                                }
                            }
                        }
                    }
                }
            }]
        elif format == 'openai':
            tools =  [
                        {
                            'type'    : 'function',
                            'function': {
                                'name'        : 'get_questions_and_answers', 
                                'description' : 'enumerate questions and answers',                  
                                'parameters'  : {
                                    'type'    : 'object',                                
                                    'properties'   : {
                                            'QAndA'  : {
                                                'type' : 'array',
                                                    'items': {
                                                        'type' : 'object',
                                                        'properties': {
                                                            'system'    : {'type': 'string'},
                                                            'user'      : {'type': 'string'},
                                                            'assistant' : {'type': 'string'},
                                                        }
                                                    }
                                                } 
                                            }                                       
                                        },
                                    },
                                }
                        ]                
        else:
            return f'Invalid format: {format}'

        # Actually make the call:
        response = await ChatCompletion.acreate(
            model        = 'gpt-4o',
            temperature  = 0.2,
            messages     = messages,
            tools        = tools,
            tool_choice  = "auto",   # <--- new param
        )

        msg = response['choices'][0]['message']
        
        # No tool calls?
        if 'tool_calls' not in msg or not msg['tool_calls']:
            # Possibly the model did not call the function. Decide on fallback:
            if retries < 4:
                return await GenQandAWithChatGPT(data, promtps, format, retries+1)
            else:
                return f'Error: no function call after multiple retries'

        # Collect all Q&A from each tool call
        args_dict = {'QAndA': []}
        for tool_call in msg['tool_calls']:
            partial_args = json.loads(tool_call['function']['arguments'])
            args_dict['QAndA'].extend(partial_args['QAndA'])

        # Possibly check if Q&A meets your minimum length
        if len(args_dict['QAndA']) < 2 and retries < 4: 
            return await GenQandAWithChatGPT(data, promtps, format, retries+1)  

        # Build your return object
        ret_dict = {data["Name"] : []}
        if format == 'json':
            for entry in args_dict['QAndA']:
                ret_dict[data["Name"]].append({
                    'answer'   : entry.get('answer', ''),
                    'question' : entry.get('question', ''),
                    'context'  : entry.get('context', '')
                })
        elif format == 'stf':
            # Just copy what was returned
            for entry in args_dict['QAndA']:
                ret_dict[data["Name"]].append(entry)
        elif format == 'stf-original':
            for entry in args_dict['QAndA']:
                ret_dict[data["Name"]].append({
                    'instruction': entry.get('instruction', ''),
                    'input'      : entry.get('input', ''),
                    'output'     : entry.get('output', '')
                })
        elif format == 'openai': 
             for entry in args_dict['QAndA']:
                if 'system' in entry.keys() and 'user' in entry.keys() and 'assistant' in entry.keys():
                    ret_dict[data["Name"]].append(OrderedDict({                        
                        'system'     : entry['system'],
                        'user'       : entry['user'], 
                        'assistant'  : entry['assistant']                      
                    }))    

        return ret_dict

    except RateLimitError:
        await async_sleep(62)
        return await GenQandAWithChatGPT(data, promtps, format, retries)

    except Exception as e:
        #AlchemyLogs().log(GetCurrentUserid(), f'{e}\n{format_exc()}', 'Exception')
        if retries < 4:
            return await GenQandAWithChatGPT(data, promtps, format, retries+1)
        return f'Error calling OpenAI API: {data["Name"]}'    

async def GenQandAWithChatGPT(data, promtps, format, retries=0): 
    #return await __GenQandAWithChatGPT(data, promtps, format, retries=retries)
    try:
        messages = GetPrompts(data, promtps, format)  

        if IsDebugEnabled():
            return loads( f'{{ "{data["Name"]}" : "Placeholder Reply for {data["Name"]}" }}' )        

        tools = [] 

        if format == 'json':
            tools =  [
                        {
                            'type'    : 'function',
                            'function': {
                                'name'        : 'get_questions_and_answers', 
                                'description' : 'enumerate questions and answers',                  
                                'parameters'  : {
                                    'type'    : 'object',                                
                                    'properties'   : {
                                            'QAndA'  : {
                                                'type' : 'array',
                                                    'items': {
                                                        'type' : 'object',
                                                        'properties': {
                                                            'question': {'type': 'string'},
                                                            'answer'  : {'type': 'string'},
                                                            'context' : {'type': 'string'},
                                                        }
                                                    }
                                                } 
                                            }                                       
                                        },
                                    },
                                }
                            ]  
        elif format == 'stf':
            tools =  [
                        {
                            'type'    : 'function',
                            'function': {
                                'name'        : 'get_questions_and_answers', 
                                'description' : 'enumerate questions and answers',                  
                                'parameters'  : {
                                    'type'    : 'object',                                
                                    'properties'   : {
                                            'QAndA'  : {
                                                'type' : 'array',
                                                    'items': {
                                                        'type' : 'object',
                                                        'properties': {
                                                            'text': {'type': 'string'}
                                                        }
                                                    }
                                                } 
                                            }                                       
                                        },
                                    },
                                }
                        ]  
        elif format == 'stf-original':
            tools =  [
                        {
                            'type'    : 'function',
                            'function': {
                                'name'        : 'get_questions_and_answers', 
                                'description' : 'enumerate questions and answers',                  
                                'parameters'  : {
                                    'type'    : 'object',                                
                                    'properties'   : {
                                            'QAndA'  : {
                                                'type' : 'array',
                                                    'items': {
                                                        'type' : 'object',
                                                        'properties': {
                                                            'instruction': {'type': 'string'},
                                                            'input'      : {'type': 'string'},
                                                            'output'     : {'type': 'string'},
                                                        }
                                                    }
                                                } 
                                            }                                       
                                        },
                                    },
                                }
                        ] 
        elif format == 'openai':
            tools =  [
                        {
                            'type'    : 'function',
                            'function': {
                                'name'        : 'get_questions_and_answers', 
                                'description' : 'enumerate questions and answers',                  
                                'parameters'  : {
                                    'type'    : 'object',                                
                                    'properties'   : {
                                            'QAndA'  : {
                                                'type' : 'array',
                                                    'items': {
                                                        'type' : 'object',
                                                        'properties': {
                                                            'system'    : {'type': 'string'},
                                                            'user'      : {'type': 'string'},
                                                            'assistant' : {'type': 'string'},
                                                        }
                                                    }
                                                } 
                                            }                                       
                                        },
                                    },
                                }
                        ]                 
               
        if len(tools) == 0: return f'Invalid format: {format}'
        
        response = await ChatCompletion.acreate(
            model           = 'gpt-4o', #'gpt-3.5-turbo' if data['Type'] == 'text' and retries==0 and len(data['Content']) < 15_000 else 'gpt-4o',            
            temperature     = 0.2,            
            messages        = messages,
            tools           = tools) 

        print(response)    
           
        msg = response['choices'][0]['message']                                          # type: ignore    
        
        #if 'tool_calls' in msg.keys():                                                  # don't check, just trigger and catch the exception               
        args_dict = loads(msg['tool_calls'][0]['function']['arguments'])                 # type: ignore   
        args_dict = {'QAndA' : []}
        for tool in msg['tool_calls']:                 
            args_dict['QAndA'].extend(loads(tool['function']['arguments'])['QAndA'])
        
        if len(args_dict['QAndA']) < 2 and retries < 4: 
            return await GenQandAWithChatGPT(data, promtps, format, retries+1)         
        print(format)
        ret_dict  = {data["Name"] : []}  
        if format == 'json':   
            for entry in args_dict['QAndA']:
                if 'question' in entry.keys() and 'answer' in entry.keys() and 'context' in entry.keys():
                    ret_dict[data["Name"]].append(OrderedDict({                        
                        'answer'   : entry['answer'],
                        'question' : entry['question'], 
                        'context'  : entry['context']                      
                    }))
                else: ret_dict[data["Name"]].append(entry)  
        elif format == 'stf':
            for entry in args_dict['QAndA']:
                ret_dict[data["Name"]].append(entry)     
        elif format == 'stf-original': 
             for entry in args_dict['QAndA']:
                if 'instruction' in entry.keys() and 'input' in entry.keys() and 'output' in entry.keys():
                    ret_dict[data["Name"]].append(OrderedDict({                        
                        'instruction' : entry['instruction'],
                        'input'       : entry['input'], 
                        'output'      : entry['output']                      
                    }))
                else: ret_dict[data["Name"]].append(entry) 
        elif format == 'openai': 
             for entry in args_dict['QAndA']:
                if 'system' in entry.keys() and 'user' in entry.keys() and 'assistant' in entry.keys():
                    ret_dict[data["Name"]].append(OrderedDict({                        
                        'system'     : entry['system'],
                        'user'       : entry['user'], 
                        'assistant'  : entry['assistant']                      
                    }))
                else: ret_dict[data["Name"]].append(entry)                             
        return ret_dict        
           
    except RateLimitError:
        await async_sleep(62)
        return await GenQandAWithChatGPT(data, promtps, format, retries)
        
    except Exception as e:    
        #AlchemyLogs().log(GetCurrentUserid(), f'{e}\n{format_exc()}', 'Exception')
        if retries < 4:            
            print('Exeption: ', str(e), format_exc(), 'Retrying...')
            return await GenQandAWithChatGPT(data, promtps, format, retries+1)            
        return f'Error calling OpenAI API: {data["Name"]}'           
    
@g_.app.route("/dataset-finetunning_gen-qanda", methods=['POST']) #type: ignore
async def Gen_Q_and_A():   
    raw_replies = await gather(*[GenQandAWithChatGPT(datum, request.json['Prompts'], request.json['format']) for datum in request.json['Files']])                      # type: ignore   
    replies     = [reply for reply in raw_replies if type(reply) == dict]                                             # type: ignore
    failed      = [failed.replace('Error calling OpenAI API:', '') for failed in raw_replies if type(failed) == str]  # type: ignore
    return jsonify({'QAndA' : replies, 'Failed' : failed})                                                            # type: ignore   